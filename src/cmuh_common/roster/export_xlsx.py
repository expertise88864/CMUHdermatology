# -*- coding: utf-8 -*-
"""匯出 Excel 值班表（openpyxl，重依賴 → 呼叫端負責 lazy 安裝）。

Sheet「值班表」：月曆式（列=週、欄=週一..週日），每格 = 日期 + R:姓名 + VS:代號。
Sheet「結算」：每位成員的平日/假日/總班/點數/帳本餘額。
"""
from __future__ import annotations

from cmuh_common.roster.export_common import (
    WD_CN, day_grid_rows, member_tally, title_text,
)
from cmuh_common.roster.model import is_weekend, week_matrix


def export(path: str, data: dict) -> None:
    from openpyxl import Workbook  # noqa: PLC0415（lazy 重依賴）
    from openpyxl.styles import Alignment

    wb = Workbook()
    _sheet_calendar(wb.active, data)
    _sheet_summary(wb.create_sheet("結算"), data)
    _sheet_day_schedule(wb.create_sheet("PGY-Clerk"), data)   # [RS-01]
    # 通用樣式：所有格自動換行、置中
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)
    wb.save(path)


def _sheet_calendar(ws, data: dict) -> None:
    from openpyxl.styles import Font, PatternFill

    ws.title = "值班表"
    ws.merge_cells("A1:G1")
    ws["A1"] = title_text(data)
    ws["A1"].font = Font(bold=True, size=14)

    for c, h in enumerate(WD_CN, start=1):
        cell = ws.cell(row=2, column=c, value=f"週{h}")
        cell.font = Font(bold=True)
        if c >= 6:                                    # 週末欄底色
            cell.fill = PatternFill("solid", fgColor="FCE4E4")

    r = data["r"]
    vs = data["vs"]
    row = 3
    for week in week_matrix(data["year"], data["month"]):
        for c, d in enumerate(week, start=1):
            if d is None:
                continue
            lines = [str(d.day)]
            rp = r["duty"].get(d)
            vp = vs["duty"].get(d)
            if rp:
                lines.append(f"R:{r['names'].get(rp, rp)}")
            if vp:
                lines.append(f"VS:{vs['names'].get(vp, vp)}")
            bp = (data.get("saturday_biopsy") or {}).get(d)   # [週六切片]
            if bp:
                lines.append(f"切:{r['names'].get(bp, bp)}")
            cell = ws.cell(row=row, column=c, value="\n".join(lines))
            if is_weekend(d) or d in data["holidays"]:
                cell.fill = PatternFill("solid", fgColor="FFF3D6")
        ws.row_dimensions[row].height = 46
        row += 1
    for c in range(1, 8):
        ws.column_dimensions[chr(64 + c)].width = 13


def _sheet_summary(ws, data: dict) -> None:
    from openpyxl.styles import Font

    heads = ("類別", "成員", "平日", "假日", "總班", "點數", "帳本")
    for c, h in enumerate(heads, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
    row = 2
    for scope in ("r", "vs"):
        block = data[scope]
        tally = member_tally(block, data["holidays"], data["params"])
        roster = set(block["members"])
        # RF-11：先列目前名單，再補「有值班但已不在名單」者（動態納入，不漏列漏點數）。
        extra = sorted(set(tally) - roster)
        for mid in list(block["members"]) + extra:
            t = tally[mid]
            name = block["names"].get(mid, mid)
            if mid in roster:
                bal_cell = round(float(block["ledger"].get(mid, 0.0)), 2)
            else:
                name = f"{name}(已離)"
                bal_cell = "—"        # 帳本已作廢，印 0.0 會被誤讀成「餘額歸零」
            ws.append([scope.upper(), name,
                       t["wd"], t["we"], t["wd"] + t["we"], t["pt"], bal_cell])
            row += 1
    for c, w in zip("ABCDEFG", (6, 10, 6, 6, 6, 6, 8)):
        ws.column_dimensions[c].width = w


def _sheet_day_schedule(ws, data: dict) -> None:
    """[RS-01] PGY/Clerk 日排班：每週一區塊（日期表頭列 + 上午/下午各一列，欄＝週一~五）。"""
    from openpyxl.styles import Font, PatternFill

    ws.title = "PGY-Clerk"
    ws.merge_cells("A1:F1")
    ws["A1"] = f"PGY / Clerk 日排班（{data['year']}/{data['month']:02d}）"
    ws["A1"].font = Font(bold=True, size=14)

    blocks = day_grid_rows(data.get("day_slots") or {}, data["year"], data["month"])
    row = 2
    for blk in blocks:
        hcell = ws.cell(row=row, column=1, value="日期")
        hcell.font = Font(bold=True)
        for c, d in enumerate(blk["weekdays"], start=2):
            v = f"{d.month}/{d.day}（{WD_CN[d.weekday()]}）" if d else ""
            cc = ws.cell(row=row, column=c, value=v)
            cc.font = Font(bold=True)
            cc.fill = PatternFill("solid", fgColor="EFEFEF")
        row += 1
        for sess, cells in blk["sessions"]:
            ws.cell(row=row, column=1, value=sess).font = Font(bold=True)
            for c, val in enumerate(cells, start=2):
                ws.cell(row=row, column=c, value=val)
            ws.row_dimensions[row].height = 40
            row += 1
        row += 1                                   # 週之間留空列
    if not blocks:
        ws["A2"] = "（本月無 PGY / Clerk 日排班）"
    ws.column_dimensions["A"].width = 8
    for c in "BCDEF":
        ws.column_dimensions[c].width = 20
